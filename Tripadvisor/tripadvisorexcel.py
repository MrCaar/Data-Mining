from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# WebDriver başlatma
url = input("URL:")
dosyaadi = input("Dosya Adi:")
driver = uc.Chrome()
driver.get(url)

# Çerezleri kabul etme
try:
    cookies_button = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Accept')]"))
    )
    cookies_button.click()
    print("Çerezler kabul edildi.")
except Exception as e:
    print("Çerez butonu bulunamadı veya zaten kabul edilmiş olabilir.")

# Excel dosyasını belirleme
excel_dosyasi = f"{dosyaadi}.xlsx"
workbook = Workbook()
sheet = workbook.active
sheet.title = "TripAdvisor Yorumlar"

# Başlıkları ekleme
sheet.append(["Kullanıcı Adı", "Tarih", "Puan", "Yorum Başlığı", "Yorum"])

# Yorumları alma fonksiyonu
def yorumlari_al():
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, "html.parser")
    divs = soup.find_all("div", {"data-automation": "tab"})

    for div in divs:
        try:
            # Kullanıcı adı
            kullaniciad_div = div.find("span", {"class": "biGQs _P fiohW fOtGX"})
            kullaniciad = kullaniciad_div.text.strip() if kullaniciad_div else "Kullanıcı Adı Bulunamadı"

            # Gönderi tarihi
            tarih_div = div.find("div", {"class": "RpeCd"})
            tarih = tarih_div.text.strip() if tarih_div else "Tarih Bulunamadı"

            # Puanı
            score_div = driver.find_element(By.XPATH, "//*[@class='UctUV d H0']/*[1]")
            score = score_div.text.strip() if score_div else "Puan Bulunamadı"
            
            # Yorum başlığı
            yorumbasligi_div = div.find("span", {"class": "yCeTE"})
            yorumbasligi = yorumbasligi_div.text.strip() if yorumbasligi_div else "Yorum Başlığı Bulunamadı"

            # Yorum kısmı
            yorum_div = div.find("span", {"class": "JguWG"})
            yorum = yorum_div.text.strip() if yorum_div else "Yorum Bulunamadı"

            # Excel'e yazma
            sheet.append([kullaniciad, tarih, score, yorumbasligi, yorum])
        except Exception as e:
            print(f"Hata oluştu: {e}")

# İlk yorumları çekme
yorumlari_al()

# Sayfa geçişleri ve yorumları çekme
while True:
    try:
        button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@aria-label='Next page']"))
        )
        button.click()
        print("Sonraki sayfaya geçildi")
        time.sleep(2.5)
        yorumlari_al()
    except Exception as e:
        print(f"Hata oldu veya sayfa bitti: {e}")
        break

driver.quit()

# Hücre genişlikleri ve yüksekliklerini ayarlama fonksiyonları
def ayarli_hucre_boyutlari(sheet, max_width=38):
    for row in sheet.iter_rows():
        row_height = 15  # Varsayılan satır yüksekliği
        for cell in row:
            if cell.value:
                text = str(cell.value)
                lines = len(text) // max_width + 1
                row_height = max(row_height, lines * 15)
                cell.alignment = Alignment(
                    wrap_text=True,  # Metni kaydır
                    vertical="center",  # Dikeyde ortala
                    horizontal="center"  # Yatayda ortala
                )
        sheet.row_dimensions[row[0].row].height = row_height

def ayarli_sutun_genislikleri(sheet, max_column_width=38):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = min(max_length + 2, max_column_width)
        sheet.column_dimensions[column].width = adjusted_width

# Genişlik ve yükseklik ayarlarını uygulama
ayarli_sutun_genislikleri(sheet)
ayarli_hucre_boyutlari(sheet)

# Excel dosyasını kaydetme
workbook.save(excel_dosyasi)
print(f"{excel_dosyasi} dosyası başarıyla kaydedildi.")
