from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
import time
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# WebDriver başlatma
driver = uc.Chrome()
url = input("URL:")
dosyaadi = input("Dosya Adi:")
driver.get(url)

# Çerezleri kabul etme
try:
    cookies_button = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Accept')]"))
    )
    cookies_button.click()
    print("Çerezler kabul edildi.")
except Exception as e:
    print("Çerez butonu bulunamadi veya zaten kabul edilmiş olabilir.")

# Excel dosyasını belirleme
excel_dosyasi = "Andulus.xlsx"

# Excel dosyasını hazırlama
excel_dosyasi = f"{dosyaadi}.xlsx"
workbook = Workbook()
sheet = workbook.active
sheet.title = "Tirpadvisor Yorumlar"

# Başlıkları ekleme
sheet.append(["Kullanıcı Adı", "Ülke", "Tarih", "Score", "Yorum Başlığı", "Yorum", "Oda Türü"])

# Yorumları alma fonksiyonu
def yorumlarıal():
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, "html.parser")
    divler = soup.find_all("div", {"data-test-target": "HR_CC_CARD"})

    for div in divler:
        try:
            # Kullanıcı adı ve ülke bilgilerini ayrı ayrı alıyoruz
            kullaniciad_div = div.find("span", {"class": "CjfFL LJbhp"})
            kullaniciad = kullaniciad_div.text.strip() if kullaniciad_div else "Kullanıcı adı bulunamadı"

            ulke_div = div.find("span", {"class": "xLwBc S2 H2 Ch d"})
            ulke = ulke_div.text.strip() if ulke_div else "Ülke bilgisi bulunamadı"

            # Diğer verileri çekiyoruz
            tarih_div = div.find("span", {"class": "iSNGb _R Me S4 H3 Cj"})
            if tarih_div:
                # <span> içeriğini hariç tutup kalan text'i al
                tarih = tarih_div.get_text(strip=True, separator=" ").replace("Date of stay:", "").strip()
                print(tarih)  # "June 2024"
            else:
                print("Kalma Tarihi Bulunamadı")

            score_div = div.find("title", id=re.compile(":lithium-"))
            if score_div:
                score_text = score_div.text.strip().split()[-1]
                try:
                    score = float(score_text.replace(",", "."))
                except ValueError:
                    score = None
            else:
                score = None

            yorumbasligi_div = div.find("span", {"class": "JbGkU Cj"})
            yorumbasligi = yorumbasligi_div.text.strip() if yorumbasligi_div else "Yorum başlığı bulunamadı"

            yorumpositive_div = div.find("span", {"class": "orRIx Ci _a C "})
            yorumpositive = yorumpositive_div.text.strip() if yorumpositive_div else "yorum bulunamadı"

            odaturu_div = div.find("span", {"class": "hHMDb _R Me"})
            if odaturu_div:
                # <span> içeriğini hariç tutup kalan text'i al
                odaturu = odaturu_div.get_text(strip=True, separator=" ").replace("Trip type: ", "").strip()
                print(odaturu)  # "June 2024"
            else:
                print("Oda Turu Bulunamadı")

            # Excel'e yazma
            sheet.append([kullaniciad, ulke, tarih, score, yorumbasligi, yorumpositive, odaturu])
        except Exception as e:
            print(f"Hata oluştu: {e}")

# Yorumları çekme ve sayfa geçişleri
try:
    #button = driver.find_element(By.CLASS_NAME, "BMQDV _F Gv wSSLS SwZTJ FGwzt ukgoS")
    #button.click()
    print("Başarıyla butona tıklandı")
    time.sleep(3)

    while True:
        yorumlarıal()
        try:
            button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@aria-label='Next page']"))
            )
            button.click()
            print("Sonraki sayfaya geçildi.")
            time.sleep(1)
        except Exception as e:
            print(f"Hata oldu veya sayfa bitti: {e}")
            break
finally:
    driver.quit()

    # Hücre genişlikleri ve yüksekliklerini ayarlama fonksiyonları
    def ayarli_hucre_boyutlari(sheet, max_width=38):
        for row in sheet.iter_rows():
            row_height = 15  # Varsayılan satır yüksekliği
            for cell in row:
                if cell.value:
                    text = str(cell.value)
                    lines = len(text) // max_width + 1
                    row_height = max(row_height, lines * 15)
                    cell.alignment = Alignment(
                        wrap_text=True,  # Metni kaydır
                        vertical="center",  # Dikeyde ortala
                        horizontal="center"  # Yatayda ortala
                    )
            sheet.row_dimensions[row[0].row].height = row_height

    def ayarli_sutun_genislikleri(sheet, max_column_width=38):
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, max_column_width)
            sheet.column_dimensions[column].width = adjusted_width

    # Genişlik ve yükseklik ayarlarını uygulama
    ayarli_sutun_genislikleri(sheet)
    ayarli_hucre_boyutlari(sheet)

    # Excel dosyasını kaydetme
    workbook.save(excel_dosyasi)
    print(f"{excel_dosyasi} dosyası başarıyla kaydedildi.")